"""Unit tests for corpus-level temporal summary."""

import openpyxl
import pytest

from src.core.corpus_summary import build_corpus_summary
from src.core.exporter import export_to_xlsx

pytestmark = pytest.mark.unit


def test_summary_groups_three_docs_in_two_years():
    docs = [
        {
            "year": "2020",
            "words_analytical": 100,
            "sent_compound_medio": 0.2,
            "leg_indice": 50,
            "sent_pct_positivo": 40,
            "sent_pct_negativo": 10,
            "term_results": {"clima": {"analytical": 2}},
            "category_results": {"CLIMA": {"analytical": 3}},
        },
        {
            "year": "2020",
            "words_analytical": 50,
            "sent_compound_medio": 0.4,
            "leg_indice": 70,
            "sent_pct_positivo": 60,
            "sent_pct_negativo": 20,
            "term_results": {"clima": {"analytical": 5}},
            "category_results": {"CLIMA": {"analytical": 7}},
        },
        {"year": "2021", "words_analytical": 30},
    ]

    summary = build_corpus_summary(docs)

    assert summary["2020"]["docs"] == 2
    assert summary["2020"]["words_analytical"] == 150
    assert summary["2020"]["sent_compound_medio"] == 0.3
    assert summary["2020"]["leg_indice_medio"] == 60
    assert summary["2020"]["terms"]["clima"] == 7
    assert summary["2020"]["categories"]["CLIMA"] == 10


def test_summary_groups_missing_year_as_no_year():
    summary = build_corpus_summary([{"year": "", "words_analytical": 12}])
    assert summary["s/ ano"]["docs"] == 1
    assert summary["s/ ano"]["words_analytical"] == 12


def test_xlsx_summary_sheet_absent_with_single_year(tmp_path):
    out = tmp_path / "r.xlsx"
    export_to_xlsx(
        [
            {
                "filename": "a.pdf",
                "year": "2020",
                "document": "",
                "president": "",
                "total_pages": 1,
                "pages_with_text": 1,
                "pages_problematic": 0,
                "ocr_pages_count": 0,
                "words_total": 10,
                "words_analytical": 10,
                "confidence": "Alto",
                "observations": "",
                "excluded_pages": [],
            }
        ],
        str(out),
    )
    assert "Síntese por Ano" not in openpyxl.load_workbook(out).sheetnames


def test_xlsx_summary_sheet_present_with_multiple_years(tmp_path):
    out = tmp_path / "r.xlsx"
    base = {
        "document": "",
        "president": "",
        "total_pages": 1,
        "pages_with_text": 1,
        "pages_problematic": 0,
        "ocr_pages_count": 0,
        "words_total": 10,
        "words_analytical": 10,
        "confidence": "Alto",
        "observations": "",
        "excluded_pages": [],
    }
    export_to_xlsx(
        [
            {"filename": "a.pdf", "year": "2020", **base},
            {"filename": "b.pdf", "year": "2021", **base},
        ],
        str(out),
    )
    wb = openpyxl.load_workbook(out)
    assert "Síntese por Ano" in wb.sheetnames
    assert wb["Síntese por Ano"].cell(row=2, column=1).value == "2020"
