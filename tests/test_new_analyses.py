"""Unit tests for n-grams, category coding and TF-IDF export."""

import openpyxl
import pytest

from src.core.analysis.base import DocumentContext
from src.core.analysis.categories import CategoryAnalyzer
from src.core.analysis.ngrams import NgramAnalyzer
from src.core.exporter import export_to_xlsx
from src.core.term_search import parse_categories, parse_input, parse_terms

pytestmark = pytest.mark.unit


# --- parser ---------------------------------------------------------------


def test_parse_categories_basic():
    cats = parse_categories('MITIGAÇÃO: clima, carbono, "efeito estufa"')
    assert cats == [
        ("MITIGAÇÃO", [("clima", False), ("carbono", False), ("efeito estufa", True)])
    ]


def test_parse_terms_skips_category_lines():
    raw = "clima\nMITIGAÇÃO: carbono, metano\ndesmatamento"
    assert parse_terms(raw) == [("clima", False), ("desmatamento", False)]


def test_parse_input_splits_both():
    terms, cats = parse_input("clima\nADAPTAÇÃO: resiliência")
    assert terms == [("clima", False)]
    assert cats[0][0] == "ADAPTAÇÃO"


def test_quoted_phrase_with_colon_is_not_category():
    terms, cats = parse_terms('"plano: metas"'), parse_categories('"plano: metas"')
    assert terms == [("plano: metas", True)]
    assert cats == []


# --- n-grams ----------------------------------------------------------------


def test_ngrams_keep_internal_stopwords():
    text = "A mudança do clima preocupa. A mudança do clima avança. " * 2
    ctx = DocumentContext("d.pdf", [text], [1], 1)
    out = NgramAnalyzer().run(ctx)
    phrases = [p for p, _n, _c in out["ngram_freq"]]
    assert "mudança do clima" in phrases


def test_ngrams_exclude_stopword_edges_and_rare():
    # "do clima" starts with a stopword; single occurrences fall below MIN_FREQUENCY.
    ctx = DocumentContext("d.pdf", ["o clima mudou hoje"], [1], 1)
    out = NgramAnalyzer().run(ctx)
    assert out["ngram_freq"] == []
    assert out["ngram_top"] == ""


# --- categories -------------------------------------------------------------


def test_category_sums_member_counts():
    cats = [("MITIGAÇÃO", [("carbono", False), ("efeito estufa", True)])]
    pages = ["carbono e efeito estufa", "carbono na capa"]
    ctx = DocumentContext("d.pdf", pages, [1], 2)  # page 2 outside corpus
    out = CategoryAnalyzer(cats).run(ctx)
    data = out["category_results"]["MITIGAÇÃO"]
    assert data["total"] == 3  # 2x carbono + 1x efeito estufa
    assert data["analytical"] == 2
    assert out["_cat_MITIGAÇÃO_total"] == 3


# --- TF-IDF export ----------------------------------------------------------


def _doc(filename, word_counts):
    return {
        "filename": filename,
        "confidence": "Alto",
        "observations": "",
        "excluded_pages": [],
        "word_counts": word_counts,
    }


def test_tfidf_sheet_ranks_distinctive_words(tmp_path):
    out = tmp_path / "r.xlsx"
    docs = [
        _doc("a.pdf", {"clima": 10, "comum": 5}),
        _doc("b.pdf", {"saude": 8, "comum": 5}),
    ]
    export_to_xlsx(docs, str(out))
    ws = openpyxl.load_workbook(out)["TF-IDF (Termos Distintivos)"]
    rows = [
        (ws.cell(row=r, column=2).value, ws.cell(row=r, column=3).value)
        for r in range(2, 5)
        if ws.cell(row=r, column=3).value
    ]
    # "comum" appears in both docs -> score 0 -> excluded.
    words = [w for _f, w in rows]
    assert "clima" in words and "saude" in words
    assert "comum" not in words


def test_tfidf_sheet_absent_with_single_doc(tmp_path):
    out = tmp_path / "r.xlsx"
    export_to_xlsx([_doc("a.pdf", {"clima": 10})], str(out))
    assert "TF-IDF (Termos Distintivos)" not in openpyxl.load_workbook(out).sheetnames
