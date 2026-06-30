"""Export results to XLSX using a dynamic, analyzer-driven column schema."""

import math
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .analysis import ColumnSpec, build_column_specs, build_default_analyzers
from .corpus_summary import build_corpus_summary, has_multiple_years
from .corpus_analysis import build_corpus_analyses
from .methodology_report import build_methodology_report, flatten_methodology_rows
from src.gui import i18n

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
TERM_HEADER_FILL = PatternFill(
    start_color="7B5CB8", end_color="7B5CB8", fill_type="solid"
)
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
_LANGUAGE = "pt"


def export_to_xlsx(
    results: List[Dict],
    output_path: str,
    column_specs: Optional[List[ColumnSpec]] = None,
    methodology_options: Optional[Dict] = None,
    language: str = "pt",
) -> None:
    """Write results to XLSX.

    ``column_specs`` is the full ordered output schema (see
    :func:`src.core.analysis.build_column_specs`). When omitted it is derived
    from the term columns present in the results, preserving backward
    compatibility with callers that do not pass an explicit schema.
    """
    global _LANGUAGE
    _LANGUAGE = i18n.normalize_language(language)
    if column_specs is None:
        column_specs = _infer_column_specs(results)

    wb = Workbook()
    ws = wb.active
    ws.title = _l("Contagem de Palavras")

    for col_idx, spec in enumerate(column_specs, start=1):
        cell = ws.cell(row=1, column=col_idx, value=spec.label)
        cell.fill = TERM_HEADER_FILL if spec.group == "term" else HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = spec.width

    ws.row_dimensions[1].height = 40
    ws.freeze_panes = "C2"

    for row_idx, result in enumerate(results, start=2):
        enriched = {"doc_id": row_idx - 1, **result}
        for col_idx, spec in enumerate(column_specs, start=1):
            value = enriched.get(spec.key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
            if row_idx % 2 == 0:
                cell.fill = ALT_ROW_FILL

    _write_excluded_sheet(wb, results)
    _write_sentiment_sheet(wb, results)
    _write_keyword_sheet(wb, results)
    _write_ngram_sheet(wb, results)
    _write_category_sheet(wb, results)
    _write_tfidf_sheet(wb, results)
    _write_kwic_sheet(wb, results)
    _write_emotion_sheet(wb, results)
    _write_geography_sheet(wb, results)
    _write_cooccurrence_sheet(wb, results)
    _write_segment_sheet(wb, results)
    _write_corpus_summary_sheet(wb, results)
    _write_corpus_analysis_sheets(wb, results)
    _write_methodology_sheet(wb, results, methodology_options)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _write_segment_sheet(wb: Workbook, results: List[Dict]) -> None:
    rows = [
        [doc_index, result.get("filename", ""), segment.get("segment_id", ""), segment.get("page_start", ""), segment.get("page_end", ""), segment.get("word_count", ""), segment.get("text", "")]
        for doc_index, result in enumerate(results, start=1)
        for segment in result.get("segments", [])
    ]
    _write_rows_sheet(
        wb,
        "Segmentos Temáticos",
        ["Nº Doc.", "Arquivo", "Segmento", "Página inicial", "Página final", "Palavras", "Texto"],
        rows,
    )


def _write_corpus_analysis_sheets(wb: Workbook, results: List[Dict]) -> None:
    analyses = build_corpus_analyses(results)
    _write_rows_sheet(
        wb,
        "Metadados Bibliográficos",
        ["Arquivo", "Título", "Autores", "Afiliações", "Ano", "Tipo", "Status"],
        [
            [
                result.get("filename", ""),
                result.get("title", ""),
                result.get("authors_display", "") or "; ".join(item.get("name", "") for item in result.get("authors", [])),
                result.get("affiliations_display", "") or "; ".join(item.get("name", "") for item in result.get("affiliations", [])),
                result.get("publication_year", "") or result.get("year", ""),
                result.get("document_type", "") or result.get("document", ""),
                result.get("metadata_status", ""),
            ]
            for result in results
        ],
    )
    entity_headers = ["Nome", "Documentos", "Documentos fracionários", "Palavras", "Ano inicial", "Ano final", "Tipos"]
    for key, title in (("authors", "Autores Consolidados"), ("institutions", "Instituições")):
        _write_rows_sheet(
            wb,
            title,
            entity_headers,
            [
                [row["name"], row["documents"], row["fractional_documents"], row["words"], row["year_start"], row["year_end"], "; ".join(row["document_types"])]
                for row in analyses["entities"][key]
            ],
        )
    _write_dict_sheet(wb, "Dispersão", analyses["dispersion"])
    _write_dict_sheet(wb, "Keyness", analyses["keyness"])
    _write_dict_sheet(wb, "Associação NPMI", analyses["cooccurrence_association"])
    similarity = analyses["similarity"]
    _write_rows_sheet(
        wb,
        "Similaridade",
        ["Documento", *similarity["labels"]],
        [[label, *similarity["matrix"][index]] for index, label in enumerate(similarity["labels"])],
    )
    _write_dict_sheet(wb, "Pares de Similaridade", similarity.get("pairs", []))
    temporal_rows = [
        {**row, "top_terms": "; ".join(row.get("top_terms", []))}
        for row in analyses["temporal_change"]
    ]
    _write_dict_sheet(wb, "Mudança Lexical", temporal_rows)
    _write_dict_sheet(wb, "Diagnóstico Sentimento", analyses["sentiment_diagnostics"])


def _write_dict_sheet(wb: Workbook, title: str, rows: List[Dict]) -> None:
    if not rows:
        return
    headers = list(rows[0])
    _write_rows_sheet(wb, title, headers, [[row.get(key, "") for key in headers] for row in rows])


def _write_rows_sheet(
    wb: Workbook, title: str, headers: List[str], rows: List[List[object]]
) -> None:
    if not rows:
        return
    ws = _detail_sheet(wb, title, headers, [max(12, min(42, len(str(header)) + 4)) for header in headers])
    for row_index, values in enumerate(rows, start=2):
        for column_index, value in enumerate(values, start=1):
            ws.cell(row=row_index, column=column_index, value=value)
        _style_detail_row(ws, row_index, len(headers))


def _infer_column_specs(results: List[Dict]) -> List[ColumnSpec]:
    """Reconstruct a schema from the analyzer outputs present in the results."""
    seen = set()
    terms: List[tuple] = []
    for r in results:
        for label in r.get("term_results", {}).keys():
            if label not in seen:
                seen.add(label)
                exact = label.startswith('"') and label.endswith('"')
                terms.append((label.strip('"') if exact else label, exact))
    # Categories are reconstructed name-only (members unknown at infer time);
    # the _cat_* values are already present in the result dicts.
    cat_names: List[str] = []
    for r in results:
        for name in r.get("category_results", {}).keys():
            if name not in cat_names:
                cat_names.append(name)
    categories = [(name, [(name, False)]) for name in cat_names]
    detect_sentiment = any("sent_n_sentencas" in r for r in results)
    detect_textmetrics = any("lex_ttr" in r for r in results)
    analyzers = build_default_analyzers(
        terms,
        detect_sentiment=detect_sentiment,
        detect_textmetrics=detect_textmetrics,
        categories=categories,
    )
    return build_column_specs(analyzers)


def _write_excluded_sheet(wb: Workbook, results: List[Dict]) -> None:
    ws2 = wb.create_sheet(_l("Páginas Excluídas"))
    headers2 = [
        "Nº Doc.",
        "Arquivo",
        "Página",
        "Motivo da Exclusão",
        "Palavras na Página",
    ]
    for col_idx, label in enumerate(headers2, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=_l(label))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 35
    ws2.column_dimensions["C"].width = 10
    ws2.column_dimensions["D"].width = 40
    ws2.column_dimensions["E"].width = 20
    ws2.row_dimensions[1].height = 28
    ws2.freeze_panes = "A2"

    detail_row = 2
    for doc_idx, result in enumerate(results, start=1):
        for page in result.get("excluded_pages", []):
            ws2.cell(row=detail_row, column=1, value=doc_idx)
            ws2.cell(row=detail_row, column=2, value=result["filename"])
            ws2.cell(row=detail_row, column=3, value=page["page_number"])
            ws2.cell(row=detail_row, column=4, value=page["exclusion_reason"])
            ws2.cell(row=detail_row, column=5, value=page["word_count"])
            for col in range(1, 6):
                ws2.cell(row=detail_row, column=col).border = THIN_BORDER
                if detail_row % 2 == 0:
                    ws2.cell(row=detail_row, column=col).fill = ALT_ROW_FILL
            detail_row += 1


def _write_sentiment_sheet(wb: Workbook, results: List[Dict]) -> None:
    """Per-sentence sentiment detail — the registry units for content analysis.

    Each row is one scored sentence with its page, compound valence and class,
    so aggregates on the main sheet are fully traceable back to the source text
    (supports Bardin's content analysis and Aguiar & Ozella's meaning nuclei).
    """
    has_detail = any(r.get("sentiment_sentences") for r in results)
    if not has_detail:
        return

    ws = wb.create_sheet(_l("Sentimento (Sentenças)"))
    headers = ["Nº Doc.", "Arquivo", "Página", "Sentença", "Compound", "Classe"]
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=_l(label))
        cell.fill = TERM_HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 90
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    detail_row = 2
    for doc_idx, result in enumerate(results, start=1):
        for sent in result.get("sentiment_sentences", []):
            ws.cell(row=detail_row, column=1, value=doc_idx)
            ws.cell(row=detail_row, column=2, value=result["filename"])
            ws.cell(row=detail_row, column=3, value=sent["page"])
            ws.cell(row=detail_row, column=4, value=sent["text"])
            ws.cell(row=detail_row, column=5, value=sent["compound"])
            ws.cell(row=detail_row, column=6, value=sent["classe"])
            ws.cell(row=detail_row, column=4).alignment = Alignment(
                vertical="center", wrap_text=True
            )
            for col in range(1, 7):
                ws.cell(row=detail_row, column=col).border = THIN_BORDER
                if detail_row % 2 == 0:
                    ws.cell(row=detail_row, column=col).fill = ALT_ROW_FILL
            detail_row += 1


def _write_keyword_sheet(wb: Workbook, results: List[Dict]) -> None:
    """Ranked keyword frequencies per document — registry units for content analysis."""
    has_detail = any(r.get("keyword_freq") for r in results)
    if not has_detail:
        return

    ws = wb.create_sheet(_l("Frequência de Palavras"))
    headers = ["Nº Doc.", "Arquivo", "Palavra", "Frequência"]
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=_l(label))
        cell.fill = TERM_HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 12
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    detail_row = 2
    for doc_idx, result in enumerate(results, start=1):
        for word, count in result.get("keyword_freq", []):
            ws.cell(row=detail_row, column=1, value=doc_idx)
            ws.cell(row=detail_row, column=2, value=result["filename"])
            ws.cell(row=detail_row, column=3, value=word)
            ws.cell(row=detail_row, column=4, value=count)
            for col in range(1, 5):
                ws.cell(row=detail_row, column=col).border = THIN_BORDER
                if detail_row % 2 == 0:
                    ws.cell(row=detail_row, column=col).fill = ALT_ROW_FILL
            detail_row += 1


def _write_kwic_sheet(wb: Workbook, results: List[Dict]) -> None:
    """Keyword-in-context concordance — the qualitative context units (Bardin)."""
    has_detail = any(r.get("kwic") for r in results)
    if not has_detail:
        return

    ws = wb.create_sheet(_l("Concordância (KWIC)"))
    headers = [
        "Nº Doc.",
        "Arquivo",
        "Página",
        "Termo",
        "Contexto à esquerda",
        "Ocorrência",
        "Contexto à direita",
    ]
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=_l(label))
        cell.fill = TERM_HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    widths = [8, 28, 8, 18, 50, 18, 50]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    right_align = Alignment(horizontal="right", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    detail_row = 2
    for doc_idx, result in enumerate(results, start=1):
        for line in result.get("kwic", []):
            ws.cell(row=detail_row, column=1, value=doc_idx)
            ws.cell(row=detail_row, column=2, value=result["filename"])
            ws.cell(row=detail_row, column=3, value=line["page"])
            ws.cell(row=detail_row, column=4, value=line["term"])
            ws.cell(
                row=detail_row, column=5, value=line["left"]
            ).alignment = right_align
            ws.cell(
                row=detail_row, column=6, value=line["keyword"]
            ).alignment = center_align
            ws.cell(row=detail_row, column=7, value=line["right"])
            for col in range(1, 8):
                ws.cell(row=detail_row, column=col).border = THIN_BORDER
                if detail_row % 2 == 0:
                    ws.cell(row=detail_row, column=col).fill = ALT_ROW_FILL
            detail_row += 1


def _detail_sheet(wb: Workbook, title: str, headers: List[str], widths: List[int]):
    """Create a styled detail sheet with a frozen header row."""
    ws = wb.create_sheet(_l(title))
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=_l(label))
        cell.fill = TERM_HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = widths[col_idx - 1]
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    return ws


def _style_detail_row(ws, row: int, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        ws.cell(row=row, column=col).border = THIN_BORDER
        if row % 2 == 0:
            ws.cell(row=row, column=col).fill = ALT_ROW_FILL


def _write_ngram_sheet(wb: Workbook, results: List[Dict]) -> None:
    """Recurring multi-word expressions — meaning-unit candidates (Bardin)."""
    if not any(r.get("ngram_freq") for r in results):
        return
    ws = _detail_sheet(
        wb, "N-gramas", ["Nº Doc.", "Arquivo", "Expressão", "N", "Frequência"],
        [8, 32, 42, 6, 12],
    )
    row = 2
    for doc_idx, result in enumerate(results, start=1):
        for phrase, n, count in result.get("ngram_freq", []):
            ws.cell(row=row, column=1, value=doc_idx)
            ws.cell(row=row, column=2, value=result["filename"])
            ws.cell(row=row, column=3, value=phrase)
            ws.cell(row=row, column=4, value=n)
            ws.cell(row=row, column=5, value=count)
            _style_detail_row(ws, row, 5)
            row += 1


def _write_category_sheet(wb: Workbook, results: List[Dict]) -> None:
    """Category coding breakdown: per-category totals and member-term counts."""
    if not any(r.get("category_results") for r in results):
        return
    ws = _detail_sheet(
        wb, "Categorias",
        ["Nº Doc.", "Arquivo", "Categoria", "Termo", "PDF Completo", "Corpus Analítico"],
        [8, 30, 24, 26, 14, 16],
    )
    bold = Font(bold=True, color="115E59")
    row = 2
    for doc_idx, result in enumerate(results, start=1):
        for name, data in result.get("category_results", {}).items():
            ws.cell(row=row, column=1, value=doc_idx)
            ws.cell(row=row, column=2, value=result["filename"])
            ws.cell(row=row, column=3, value=name).font = bold
            ws.cell(row=row, column=4, value="(total da categoria)").font = bold
            ws.cell(row=row, column=5, value=data["total"]).font = bold
            ws.cell(row=row, column=6, value=data["analytical"]).font = bold
            _style_detail_row(ws, row, 6)
            row += 1
            for label, counts in data.get("members", {}).items():
                ws.cell(row=row, column=1, value=doc_idx)
                ws.cell(row=row, column=2, value=result["filename"])
                ws.cell(row=row, column=3, value=name)
                ws.cell(row=row, column=4, value=label)
                ws.cell(row=row, column=5, value=counts["total"])
                ws.cell(row=row, column=6, value=counts["analytical"])
                _style_detail_row(ws, row, 6)
                row += 1


TFIDF_TOP_N = 15


def _write_tfidf_sheet(wb: Workbook, results: List[Dict]) -> None:
    """Distinctive words per document via TF-IDF (Salton & Buckley, 1988).

    tf-idf(w, d) = count(w, d) * ln(N / df(w)). Words present in every document
    score zero (they are not distinctive). Requires at least two documents with
    word counts.
    """
    docs = [r for r in results if r.get("word_counts")]
    n_docs = len(docs)
    if n_docs < 2:
        return

    df: Dict[str, int] = {}
    for r in docs:
        for word in r["word_counts"]:
            df[word] = df.get(word, 0) + 1

    ws = _detail_sheet(
        wb, "TF-IDF (Termos Distintivos)",
        ["Nº Doc.", "Arquivo", "Palavra", "Frequência", "TF-IDF"],
        [8, 32, 28, 12, 12],
    )
    row = 2
    for doc_idx, result in enumerate(results, start=1):
        counts = result.get("word_counts")
        if not counts:
            continue
        scored = [
            (word, count, round(count * math.log(n_docs / df[word]), 3))
            for word, count in counts.items()
        ]
        scored = [s for s in scored if s[2] > 0]
        scored.sort(key=lambda s: s[2], reverse=True)
        for word, count, score in scored[:TFIDF_TOP_N]:
            ws.cell(row=row, column=1, value=doc_idx)
            ws.cell(row=row, column=2, value=result["filename"])
            ws.cell(row=row, column=3, value=word)
            ws.cell(row=row, column=4, value=count)
            ws.cell(row=row, column=5, value=score)
            _style_detail_row(ws, row, 5)
            row += 1


def _write_geography_sheet(wb: Workbook, results: List[Dict]) -> None:
    if not any(r.get("geo_mentions") for r in results):
        return
    ws = _detail_sheet(
        wb,
        "Menções Territoriais",
        ["Nº Doc.", "Arquivo", "Local", "Tipo", "UF", "Contagem"],
        [8, 32, 30, 16, 8, 12],
    )
    row = 2
    for doc_idx, result in enumerate(results, start=1):
        for name, place_type, uf, count in result.get("geo_mentions", []):
            ws.cell(row=row, column=1, value=doc_idx)
            ws.cell(row=row, column=2, value=result["filename"])
            ws.cell(row=row, column=3, value=name)
            ws.cell(row=row, column=4, value=place_type)
            ws.cell(row=row, column=5, value=uf)
            ws.cell(row=row, column=6, value=count)
            _style_detail_row(ws, row, 6)
            row += 1


def _write_emotion_sheet(wb: Workbook, results: List[Dict]) -> None:
    if not any(r.get("emotion_words") for r in results):
        return
    ws = _detail_sheet(
        wb,
        "Emoções (Palavras)",
        ["Nº Doc.", "Arquivo", "Emoção", "Palavra", "Contagem"],
        [8, 32, 20, 28, 12],
    )
    row = 2
    for doc_idx, result in enumerate(results, start=1):
        for emotion, words in result.get("emotion_words", {}).items():
            for word, count in words:
                ws.cell(row=row, column=1, value=doc_idx)
                ws.cell(row=row, column=2, value=result["filename"])
                ws.cell(row=row, column=3, value=emotion)
                ws.cell(row=row, column=4, value=word)
                ws.cell(row=row, column=5, value=count)
                _style_detail_row(ws, row, 5)
                row += 1


def _write_cooccurrence_sheet(wb: Workbook, results: List[Dict]) -> None:
    if not any(r.get("cooccurrence") for r in results):
        return
    ws = _detail_sheet(
        wb,
        "Co-ocorrência",
        ["Nº Doc.", "Arquivo", "Termo A", "Termo B", "Sentenças"],
        [8, 32, 28, 28, 12],
    )
    row = 2
    for doc_idx, result in enumerate(results, start=1):
        for term_a, term_b, count in result.get("cooccurrence", []):
            ws.cell(row=row, column=1, value=doc_idx)
            ws.cell(row=row, column=2, value=result["filename"])
            ws.cell(row=row, column=3, value=term_a)
            ws.cell(row=row, column=4, value=term_b)
            ws.cell(row=row, column=5, value=count)
            _style_detail_row(ws, row, 5)
            row += 1


def _write_corpus_summary_sheet(wb: Workbook, results: List[Dict]) -> None:
    if not has_multiple_years(results):
        return
    summary = build_corpus_summary(results)
    term_labels = []
    category_names = []
    for row in summary.values():
        for label in row["terms"]:
            if label not in term_labels:
                term_labels.append(label)
        for name in row["categories"]:
            if name not in category_names:
                category_names.append(name)

    headers = [
        "Ano",
        "Nº Docs",
        "Palavras (Corpus)",
        "Sentimento médio",
        "Legibilidade média",
        "% positivas",
        "% negativas",
    ]
    headers += [f"Termo: {label}" for label in term_labels]
    headers += [f"Categoria: {name}" for name in category_names]
    widths = [10, 10, 18, 16, 18, 12, 12] + [18] * (
        len(term_labels) + len(category_names)
    )
    ws = _detail_sheet(wb, "Síntese por Ano", headers, widths)
    row_idx = 2
    for year, data in summary.items():
        values = [
            year,
            data["docs"],
            data["words_analytical"],
            data["sent_compound_medio"],
            data["leg_indice_medio"],
            data["sent_pct_positivo"],
            data["sent_pct_negativo"],
        ]
        values += [data["terms"].get(label, 0) for label in term_labels]
        values += [data["categories"].get(name, 0) for name in category_names]
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        _style_detail_row(ws, row_idx, len(values))
        row_idx += 1


def _write_methodology_sheet(
    wb: Workbook, results: List[Dict], methodology_options: Optional[Dict]
) -> None:
    report = build_methodology_report(results, methodology_options)
    ws = _detail_sheet(wb, "Metodologia", ["Item", "Valor"], [30, 100])
    for row_idx, (key, value) in enumerate(flatten_methodology_rows(report, _LANGUAGE), start=2):
        ws.cell(row=row_idx, column=1, value=key)
        ws.cell(row=row_idx, column=2, value=value)
        ws.cell(row=row_idx, column=2).alignment = Alignment(
            vertical="top", wrap_text=True
        )
        _style_detail_row(ws, row_idx, 2)


def _l(text: str) -> str:
    return i18n.label(text, _LANGUAGE)
